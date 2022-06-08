# -*- coding:utf8 -*- #
# -----------------------------------------------------------------------------------
# ProjectName:   BMP_request
# FileName:      read_excel
# Author:        ZJY
# Datetime:      2022/6/2 10:45
# Description：
# -----------------------------------------------------------------------------------
import time

import openpyxl
from openpyxl.styles import Font  # 导入字体模块

from tool import read_ini


class ReadExcel:

    def __init__(self, section, option, sheet):
        # 读取文件路径
        self.read_ini = read_ini.ReadIni()
        self.file_path = self.read_ini.get_all_path(section, option)
        # print(file_path)
        # 实例化对象
        self.wb = openpyxl.load_workbook(self.file_path)
        self.ws = self.wb[sheet]
        # # 去掉表格中的空格和换行符
        # for row in self.ws:
        #     for cell in row:
        #         old = cell.value
        #         if old:
        #             cell.value = old.replace(' ', '').replace('\n', '')

    # 编号	模块名称	接口名称	用例标题	用例等级	请求的方法	请求的路径	传参的类型	用例数据	期望数据	执行的结果

    # 获取excel用例条数
    def get_row(self):
        return self.ws.max_row

    # 获取excel最大行数
    def get_col(self):
        return self.ws.max_column

    # 读取表格内容
    def get_cell_data(self, column, row):
        return self.ws[column + str(row)].value

    # 获取用例编号
    def get_case_order(self, row):
        return self.get_cell_data('A', row)

    # 获取模块名称
    def get_case_module(self, row):
        return self.get_cell_data('B', row)

    # 获取接口名称
    def get_case_interface(self, row):
        return self.get_cell_data('C', row)

    # 获取用例标题
    def get_case_title(self, row):
        return self.get_cell_data('D', row)

    # 获取用例等级
    def get_case_grade(self, row):
        return self.get_cell_data('E', row)

    # 获取请求方法
    def get_request_method(self, row):
        return self.get_cell_data('F', row)

    # 获取请求路径
    def get_request_path(self, row):
        return self.get_cell_data('G', row)

    # 获取传参类型
    def get_params_type(self, row):
        if self.get_cell_data('H', row):
            return eval(self.get_cell_data('H', row))

    # 获取用例数据
    def get_case_data(self, row):
        if self.get_cell_data('I', row):
            return eval(self.get_cell_data('I', row).replace(' ', '').replace('\n', ''))  # 去掉空格和换行符

    # 获取期望数据
    def get_expect_data(self, row):
        return eval(self.get_cell_data('J', row).replace(' ', '').replace('\n', ''))  # 去掉空格和换行符

    # 获取SQL语句类型
    def get_sql_type(self, row):
        if self.get_cell_data('k', row):
            return eval(self.get_cell_data('k', row))

    # 获取SQL语句
    def get_sql(self, row):
        if self.get_cell_data('l', row):
            return eval(self.get_cell_data('l', row))
        # if sql_key:
        #     # 1:读取sql.json文件
        #     sql_data_json = get_json_data(self.read_ini.read_all_path('path', 'json_sql_path'))
        #     # 2：获取模块名称
        #     module_name = self.get_case_module(row)
        #     # 3：获取接口名称
        #     interface_name = self.get_case_interface(row)
        #     return sql_data_json[module_name][interface_name][sql_key]

    # 获取更新的key
    def get_new_key(self, row):
        if self.get_cell_data('m', row):
            return eval(self.get_cell_data('m', row))

    # 获取想要的数据到列表
    def get_case_data_list(self):
        list1 = []
        for row in range(2, self.get_row() + 1):
            request_method = self.get_request_method(row)
            url = self.read_ini.get_after_path('host', 'bmp_host') + self.get_request_path(row)
            params_types = self.get_params_type(row)
            data_case = self.get_case_data(row)
            expect_data = self.get_expect_data(row)
            sql_types = self.get_sql_type(row)
            sql = self.get_sql(row)
            new_key = self.get_new_key(row)
            module = self.get_case_module(row)
            interface = self.get_case_interface(row)
            title = self.get_case_title(row)
            list1.append([module, interface, title, request_method, url, params_types, data_case,
                          expect_data, row, sql_types, sql, new_key])
        return list1

    # 写入结果到excel
    def save_result(self, result, row):
        self.ws['n' + str(row)] = result + time.strftime("%Y-%m-%d %H:%M:%S")
        if result == '失败':
            # 失败的结果红色标注
            self.ws['n' + str(row)].font = Font(color='ff0033')
        else:
            self.ws['n' + str(row)].font = Font(color='000000')
        self.wb.save(self.file_path)


if __name__ == '__main__':
    read_excel = ReadExcel('path', 'excel_path', 'Sheet1')
    print(ReadExcel('path', 'excel_path', 'Sheet1').get_case_data_list())
    key1 = read_excel.get_params_type(31)
    key2 = read_excel.get_new_key(26)
    print(type(key1),key1, key2)
    print(read_excel.open_case_data(9))
    read_excel.open_case_data(31)
    print(json_data, params_data)
    # print(key1[0])
